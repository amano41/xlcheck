import os
import sys
from pathlib import Path
from typing import Union, cast

import openpyxl as xl
from openpyxl import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from .answer import Answer

PATH_TYPE = Union[str, bytes, os.PathLike]


def usage():
    cmd = Path(__file__).name
    print(f"Usage: {cmd} <workbook> <answer>")
    print(f"       {cmd} <directory> <answer>")


def main():
    if len(sys.argv) != 3:
        usage()
        exit()

    target_path = Path(sys.argv[1])
    answer_path = Path(sys.argv[2])

    # 結果出力時のヘッダ
    header = "\t".join(("Sheet", "Cell", "Formula", "Result"))

    # 模範解答
    answer = Answer(answer_path)

    # 採点対象がファイルの場合
    if target_path.is_file():
        print(header)
        for r in check_file(target_path, answer):
            print("\t".join(map(str, r)))

    # 採点対象がディレクトリの場合
    elif target_path.is_dir():
        for target_file in target_path.glob("*.xlsx"):
            print(target_file)
            b = xl.load_workbook(str(target_file))

            result = check(b, answer)

            output_file = target_file.with_suffix(".tsv")
            with output_file.open("w", encoding="utf-8", newline="\n") as f:
                f.write(header + "\n")
                for r in result:
                    f.write("\t".join(map(str, r)) + "\n")

    # 読み込めない場合はエラー
    else:
        print(f"Error: No such file or directory: {target_path}", file=sys.stderr)


def check_file(workbook_path: PATH_TYPE, answer: Answer) -> list[tuple[str, str, str, bool]]:
    b = xl.load_workbook(str(workbook_path))
    return check(b, answer)


def check(workbook: Workbook, answer: Answer) -> list[tuple[str, str, str, bool]]:
    result = []

    for s in answer.sheets():
        worksheet = _find_worksheet(workbook, s)

        # ワークシート内の配列数式・スピルを調べる
        array_formulae = {}
        for coord, ref in worksheet.array_formulae.items():
            rng = CellRange(str(ref))
            array_formulae[rng] = worksheet[str(coord)].value.text

        for c in answer.cells(s):
            # 配列数式・スピルの一部ならば基準セルの値を使う
            rng = CellRange(c)
            for array in array_formulae:
                if rng.issubset(array):
                    v = array_formulae[array]
                    break
            # 配列数式・スピルでなければセルの値をそのまま使う
            else:
                if worksheet[c].value is None:
                    v = ""
                else:
                    v = str(worksheet[c].value)

            # 解答に一致するかチェック
            r = answer.match(s, c, v)
            result.append((s, c, v, r))

    return result


def _find_worksheet(workbook: Workbook, sheetname: str) -> Worksheet:
    # 名前が一致すればそのまま返す
    if sheetname in workbook:
        return cast(Worksheet, workbook[sheetname])

    # 大文字・小文字を区別せずにチェック
    sheetname = sheetname.lower()
    for name in workbook.sheetnames:
        if name.lower() == sheetname:
            return cast(Worksheet, workbook[name])

    # 見つからなければブランクシート
    return cast(Worksheet, workbook.create_sheet())


if __name__ == "__main__":
    main()
